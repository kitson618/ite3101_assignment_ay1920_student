from typing import List, Dict

#Kitson
import pandas as pd
from pathlib import Path
from sql_to_python.excel.tool import generate_excel
from sql_to_python.file.file_helper import unzip_to_folder, list_all_files, save_to_file, delete_folder
from sql_to_python.filter.logic import get_result_from_file, get_valid_and_invalid_exercise_lists
from openpyxl import load_workbook

# https://pypi.org/project/html5print/
from html5print import HTMLBeautifier
from sql_to_python.filter.logic import generate_data_list
from sql_to_python.excel.tool import get_excel_sheetnames

def generate_html_template(headers: List[str]) -> str:
    tab = "".join(map(lambda x: f"<button class=\"tablinks\" onclick=\"openQuestion(event, '{x}')\">{x}</button>", headers))
    return f"""<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="stylesheet" type="text/css" href="../sql_to_python/html/style.css">
<script type="text/javascript" src="../sql_to_python/html/tab.js"></script>
</head>
<body>
<h2>SQL - Python Translator</h2>
<!-- tab -->
<div class="tab">
{tab}
</div>
<!-- table content -->
###table_body###
</body>
</html>"""

def generate_all_th(rows: List[str]) -> str:
    tds = "".join(map(lambda x: f"<th>{x}</th>", rows))
    return f"<tr>{tds}</tr>"


def generate_all_tr(rows: List[List[str]]) -> str:
    return "".join(map(lambda x: generate_tr(x), rows))


def generate_tr(rows: List[str]) -> str:
    tds = "".join(map(lambda x: f"<td>{x}</td>", rows))
    return f"<tr>{tds}</tr>"


#def generate_html_report(file_path_name: Path, results: List[Dict[str, str]]) -> str:

def generate_html_report(file_path_name: Path) -> str:

    tbl = ""
    sheets = get_excel_sheetnames(file_path_name)

    for i in sheets:
        df = pd.read_excel(file_path_name, sheet_name=i, dtype=str)
        thead = generate_all_th(df.columns)
        trs = generate_all_tr(df.values.tolist())
        tbl += f"<div id=\"{i}\" class=\"tabcontent\"><table>{thead}{trs}</table></div>"
    template = generate_html_template(sheets)
    return HTMLBeautifier.beautify(template.replace("###table_body###", tbl))


file = Path("../../out/output.xlsx")
out_folder = Path("../../out/")
html_report = out_folder.joinpath("result.html")
html_report_content = generate_html_report(file)
save_to_file(html_report, html_report_content)


